import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from scipy import stats
from scipy.io import loadmat


m = loadmat('D:\codespace\ICN_tensor_GCN\sub_ICN_tensor_gcn\\1dianxian0.80.010.001.mat')
keysm = list(m.keys())
print(keysm)
fea1 = m['feas1']
fea2 = m['feas2']
fea3 = m['feas3']
fea4 = m['feas4']
fea5 = m['feas5']
fea6 = m['feas6']
fea7 = m['feas7']
fea8 = m['feas8']

# print(fea1[0][0])
# print(fea1[1][0].shape)
labels = m['label'][0]
my_list = list(labels)
select_sub2 = []
select_sub1 = []
target_element0 = 0

for i in range(len(my_list)):
    if my_list[i] == target_element0:
        select_sub1.append(i)
target_element1 = 1

for i in range(len(my_list)):
    if my_list[i] == target_element1:
        select_sub2.append(i)

select_fea11 = np.mean(fea1[select_sub1, :, :, :], axis=(0, 1))
select_fea21 = np.mean(fea1[select_sub2, :, :, :], axis=(0, 1))
print('DMN', np.mean((np.abs(np.corrcoef(select_fea11))), axis=(0, 1)))
print('DMN', np.mean((np.abs(np.corrcoef(select_fea21))), axis=(0, 1)))

select_fea12 = np.mean(fea2[select_sub1, :, :, :], axis=(0, 1))
select_fea22 = np.mean(fea2[select_sub2, :, :, :], axis=(0, 1))
print('FPN', np.mean((np.abs(np.corrcoef(select_fea12))), axis=(0, 1)))
print('FPN', np.mean((np.abs(np.corrcoef(select_fea22))), axis=(0, 1)))

select_fea13 = np.mean(fea3[select_sub1, :, :, :], axis=(0, 1))
select_fea23 = np.mean(fea3[select_sub2, :, :, :], axis=(0, 1))
print('SAN', np.mean((np.abs(np.corrcoef(select_fea13))), axis=(0, 1)))
print('SAN', np.mean((np.abs(np.corrcoef(select_fea23))), axis=(0, 1)))

select_fea14 = np.mean(fea4[select_sub1, :, :, :], axis=(0, 1))
select_fea24 = np.mean(fea4[select_sub2, :, :, :], axis=(0, 1))
print('ATN',np.mean((np.abs(np.corrcoef(select_fea14))), axis=(0, 1)))
print('ATN',np.mean((np.abs(np.corrcoef(select_fea24))), axis=(0, 1)))

select_fea15 = np.mean(fea5[select_sub1, :, :, :], axis=(0, 1))
select_fea25 = np.mean(fea5[select_sub2, :, :, :], axis=(0, 1))
print('SMN',np.mean((np.abs(np.corrcoef(select_fea15))), axis=(0, 1)))
print('SMN',np.mean((np.abs(np.corrcoef(select_fea25))), axis=(0, 1)))

select_fea16 = np.mean(fea6[select_sub1, :, :, :], axis=(0, 1))
select_fea26 = np.mean(fea6[select_sub2, :, :, :], axis=(0, 1))
print('VN',np.mean((np.abs(np.corrcoef(select_fea16))), axis=(0, 1)))
print('VN',np.mean((np.abs(np.corrcoef(select_fea26))), axis=(0, 1)))

select_fea17 = np.mean(fea7[select_sub1, :, :, :], axis=(0, 1))
select_fea27 = np.mean(fea7[select_sub2, :, :, :], axis=(0, 1))
print('AN',np.mean((np.abs(np.corrcoef(select_fea17))), axis=(0, 1)))
print('AN',np.mean((np.abs(np.corrcoef(select_fea27))), axis=(0, 1)))

select_fea18 = np.mean(fea8[select_sub1, :, :, :], axis=(0, 1))
select_fea28 = np.mean(fea8[select_sub2, :, :, :], axis=(0, 1))
print('ON',np.mean((np.abs(np.corrcoef(select_fea18))), axis=(0, 1)))
print('ON',np.mean((np.abs(np.corrcoef(select_fea28))), axis=(0, 1)))


# Spatio-temporal variation rate
select_fea11 = fea1[select_sub1, :, :, :]
select_fea21 = fea1[select_sub2, :, :, :]

select_fea12 = fea2[select_sub1, :, :, :]
select_fea22 = fea2[select_sub2, :, :, :]

select_fea13 = fea3[select_sub1, :, :, :]
select_fea23 = fea3[select_sub2, :, :, :]

select_fea14 = fea4[select_sub1, :, :, :]
select_fea24 = fea4[select_sub2, :, :, :]

select_fea15 = fea5[select_sub1, :, :, :]
select_fea25 = fea5[select_sub2, :, :, :]

select_fea16 = fea6[select_sub1, :, :, :]
select_fea26 = fea6[select_sub2, :, :, :]

select_fea17 = fea7[select_sub1, :, :, :]
select_fea27 = fea7[select_sub2, :, :, :]

select_fea18 = fea8[select_sub1, :, :, :]
select_fea28 = fea8[select_sub2, :, :, :]
NC = np.concatenate(
    (select_fea11, select_fea12, select_fea13, select_fea14, select_fea15, select_fea16, select_fea17, select_fea18),
    axis=2)
ILL = np.concatenate(
    (select_fea21, select_fea22, select_fea23, select_fea24, select_fea25, select_fea26, select_fea27, select_fea28),
    axis=2)

# for yu in np.arange(0.0, 0.1, 0.001):
yu = 0
dy_net_all1 = []
for i in range(NC.shape[0]):
    dy_net = []
    for j in range(NC.shape[1]):
        net = np.abs(np.corrcoef(NC[i][j]))
        net[net <= yu] = 0
        net = np.sum(net)
        # print(net)
        dy_net.append(net)
    dy_net = np.array(dy_net)
    var1 = np.var(dy_net, 0)
    dy_net_all1.append(var1)
dy_net_all1 = np.array(dy_net_all1)
# print(list(dy_net_all1))
print(np.mean(dy_net_all1))

dy_net_all2 = []
for i in range(ILL.shape[0]):
    dy_net = []
    for j in range(ILL.shape[1]):
        net = np.abs(np.corrcoef(ILL[i][j]))
        net[net <= yu] = 0
        net = np.sum(net)
        dy_net.append(net)
    dy_net = np.array(dy_net)
    var1 = np.var(dy_net, 0)
    dy_net_all2.append(var1)
dy_net_all2 = np.array(dy_net_all2)
# print(list(dy_net_all2))
print(np.mean(dy_net_all2))
t_statistic, p_value = stats.ttest_ind(dy_net_all1, dy_net_all2)
if p_value < 0.05:
    print("t:", t_statistic)
    print("p:", p_value)


def write_arrays_to_excel(array1, array2, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 写入第一个数组到第一列
    for i, value in enumerate(array1):
        cell = sheet.cell(row=i + 1, column=1)
        cell.value = value

    # 写入第二个数组到第二列
    for i, value in enumerate(array2):
        cell = sheet.cell(row=i + 1, column=2)
        cell.value = value

    # 保存 Excel 文件
    workbook.save(file_path)



